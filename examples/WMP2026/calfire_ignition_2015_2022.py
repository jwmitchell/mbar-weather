# 2025 data set of outages. Will check FCS outages versus others. Key
# variables are temperature (max) wind gust (max), and humidity(min).
# Window should be < 1 hour

# Data provided in last year's PG&E WMP includes all historic large calfire ignitions in PG&E service area.
# Get weather station data within 5 miles and within 4 hours of ignition 


import sys
sys.path.append("..")
from datetime import datetime
import pytz
import logging
from openpyxl import load_workbook
import xlrd  # pip install
from importlib import import_module
sys.path.append("..")

parse = argparse.ArgumentParser()
parse.add_argument('-u','--utility',choices=['PGE','SCE','SDGE'],required=True,help='utility=PGE,SCE,SDGE')
parse.add_argument('-f','--file',help='file=configuration file')
program_args=parse.parse_args()

import weather_config

if program_args.file is not None:
    weather_config.init(program_args.file)

import weather_utils

logging.basicConfig(level=weather_config.config['Default']['LOG_LEVEL'])

xl_data = weather_config.config['PGE']['XL_DATA_FILE']
weather_db = weather_config.config['PGE']['WEATHER_DB']
xl_input_sheet = weather_config.config['PGE']['XL_INPUT_SHEET']
xl_output_sheet = weather_config.config['PGE']['XL_OUTPUT_SHEET']
ttpl_raw = weather_config.config['PGE']['TIME_WINDOWS']
gtpl_raw = weather_config.config['PGE']['DISTANCE_WINDOWS']
xl_datetime_column = weather_config.config['PGE']['XL_DATETIME_COLUMN']
xl_latitude_column = weather_config.config['PGE']['XL_LATITUDE_COLUMN']
xl_longitude_column = weather_config.config['PGE']['XL_LONGITUDE_COLUMN']

tlst = ttpl_raw.split(',')
ttpl = (int(tlst[0]),int(tlst[1]))
glst = gtpl_raw.split(',')
gtpl = (int(glst[0]),int(glst[1]))

try: 
    cfigndb = weather_utils.WeatherDB(weather_db)   
except:
    cfigndb = weather_utils.WeatherDB.create(weather_db)

logging.info('Opening workbook ' + xl_data)
wbk = load_workbook(filename=xl_data)

sht_all = wbk[xl_input_sheet]

try:
    sht_wind = wbk[xl_output_sheet]
except KeyError:
    wbk.create_sheet(xl_output_sheet)
    sht_wind = wbk[xl_output_sheet]
    
frow = int(weather_config.config['PGE']['FIRST_ROW'])
lrow = int(weather_config.config['PGE']['LAST_ROW'])
#weather_station_cell = weather_config.config['PGE']['WS'] No weather station

for irow in range(frow,lrow+1):
    
    # Copy row

    srow = str(irow)
    logging.info("Processing line " + srow)
    for icol in range(len(sht_all[srow])):
        sht_wind.cell(row=irow,column=icol+1).value = sht_all[irow][icol].value

    # Convert times to UTC for synoptic run. PG&E has multiple date/time formats
    # in their data set:
    # Case 1   B: Excel datetime    C: Excel datetime
    # Case 2   B: Excel date        C: Excel datetime
    # Case 3   B: Excel datetime    C: Excel time
    # Case 4   B: Excel date        C: Excel time
    # Case 5   B: Excel datetime    C: HH:MM
    # Case 6   B: DD/MM/YYYY        C: HH:MM  - Adding this for additional fires

#    tcelld = "D"+ srow
#    tcellt = "F"+ srow
#    tmxld = sht_wind[tcelld].value
#    tmxlt = sht_wind[tcellt].value

    

    is_xl_format = False
    tmxl = 0

    try:

        tmpydt = sht_wind[xl_ignition_column+srow].value
        
    except TypeError:
        logging.warning("Exiting on error. Saving workbook " + xl_data)
        wbk.save(xl_data)
        raise

    tmlocal = pytz.timezone("America/Los_Angeles").localize(tmpydt)
    tmutc = tmlocal.astimezone(pytz.utc)

    zigtime = weather_utils.TimeUtils(tmutc)
#    lat = sht_wind[XL_LATITUDE_COLUMN+srow].value
#    lon = sht_wind[XL_LONGITUDE_COLUMN+srow].value

    # Get list of max wind data based on time and space windows
    # Output is m x n list of  [station_id,time,distance,maximum gust,count]

    ttpl = (ttpl,) if isinstance(ttpl,int) else ttpl
    gtpl = (gtpl,) if isinstance(gtpl,int) else gtpl
    
    try:
#        ws_values = weather_utils.get_observations_by_stid_datetime(stid,firstdt,lastdt,db_object)
        max_gusts = weather_utils.get_max_gust(lat,lon,zigtime,ttpl,gtpl,cfigndb)
    except:
        logging.warning("Exiting on error. Saving workbook " + xl_data)
        wbk.save(xl_data)
        raise
    
    # Write output to cells
    
    icell = int(weather_config.config['SCE']['FREE_CELL'])
    for it in max_gusts:
        for ig in it:
            for val in ig:
                sht_wind.cell(row=irow,column=icell).value = val
                icell += 1

logging.info("Complete. Saving workbook " + xl_data)
wbk.save(xl_data)
