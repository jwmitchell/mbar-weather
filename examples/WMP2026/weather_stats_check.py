# Portions of this code have been generated by ChatGPT 4.0
# Transcript reference: https://chatgpt.com/canvas/shared/680e661521488191b094531764a81283
# Wind Speed Distribution Analysis Script with T-Tests for Averages.
# This tests two distributions.
# Argument parsing, Excel features added from standard wind libraries
# Joseph W. Mitchell   April 2025

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.stats import chi2, ks_2samp, ttest_ind
import argparse
from openpyxl import load_workbook
import openpyxl
import xlrd  # pip install

parse = argparse.ArgumentParser()
parse.add_argument('-s','--sheet',required=True,help='Name of Sheet in Excel')
parse.add_argument('-o','--output', help = 'Name of Output Sheet in Excel')
parse.add_argument('-f','--file',required=True,help='Excel File')
parse.add_argument('-c','--c1',help='Column 1')
parse.add_argument('-d','--c2',help='Column 2')
parse.add_argument('-l','--labels',help='Row Labels')
parse.add_argument('-r','--row',help='Row 1')
parse.add_argument('-n','--n',help='Number of Points')
parse.add_argument('-z','--outcell',help="Output Cell")

program_args=parse.parse_args()

#xl_default = [Column 1, Column 2, Labels, Row, Number] - When entering on command line becomes tedious
xl_default = {"c1":"E", "c2":"Q", "label":"D", "row":25, "n":6}

xl_file = program_args.file
xl_insheet = program_args.sheet
if program_args.output != None:
    xl_outsheet = program_args.output
else:
    xl_outsheet = 'Stats'

if program_args.c1 == None or program_args.c2 == None or program_args.label == None or program_args.row == None or program_args.n == None:
    xl_data = xl_default
else:
    xl_data = {"c1": program_args.c1,"c2": program_args.c2,"label": program_args.label,"r1": program_args.r1,"n": program_args.n}
    
def pull_data_from_excel(xlfile,xlinsheet,xldata):
    wbk = load_workbook(filename=xlfile)
    sht_in = wbk[xlinsheet]
    labels = []
    dist1 = []
    dist2 = []
    row0 = xldata['row']
    for i in range(row0,row0+xldata['n']):
        labels.append(sht_in[xldata['label']+str(i)].value)
        dist1.append(sht_in[xldata['c1']+str(i)].value)
        dist2.append(sht_in[xldata['c2']+str(i)].value)
    wbk.close()      
    return(labels,dist1,dist2)
    
    
# --- Input your compressed distributions (counts) ---
#dist_a = [3, 16, 16, 10, 5, 5]
#dist_b = [22, 130, 86, 25, 9, 7]

labels,dist_a,dist_b = pull_data_from_excel(xl_file,xl_insheet,xl_data)

# --- G-test per bin ---
def g_test_per_bin(dist_a, dist_b):
    observed = np.vstack([dist_a, dist_b])
    row_totals = observed.sum(axis=1, keepdims=True)
    col_totals = observed.sum(axis=0, keepdims=True)
    grand_total = observed.sum()
    expected = row_totals @ col_totals / grand_total
    with np.errstate(divide='ignore', invalid='ignore'):
        g_per_cell = 2 * observed * np.log(np.where(observed == 0, 1, observed / expected))
        g_per_cell = np.nan_to_num(g_per_cell)
    g_stat_per_bin = g_per_cell.sum(axis=0)
    p_val_per_bin = chi2.sf(g_stat_per_bin, df=1)
    bins = np.arange(1, len(dist_a) + 1)
    return list(zip(bins, g_stat_per_bin, p_val_per_bin))

# --- Chi-squared test per bin ---
def chi2_test_per_bin(dist_a, dist_b):
    observed = np.vstack([dist_a, dist_b])
    row_totals = observed.sum(axis=1, keepdims=True)
    col_totals = observed.sum(axis=0, keepdims=True)
    grand_total = observed.sum()
    expected = row_totals @ col_totals / grand_total
    with np.errstate(divide='ignore', invalid='ignore'):
        chi2_per_cell = (observed - expected) ** 2 / np.where(expected == 0, 1, expected)
        chi2_per_cell = np.nan_to_num(chi2_per_cell)
    chi2_stat_per_bin = chi2_per_cell.sum(axis=0)
    p_val_per_bin = chi2.sf(chi2_stat_per_bin, df=1)
    bins = np.arange(1, len(dist_a) + 1)
    return list(zip(bins, chi2_stat_per_bin, p_val_per_bin))

# --- Expand counts into samples for KS test ---
def expand_to_samples(dist):
    return np.concatenate([np.full(count, bin_idx) for bin_idx, count in enumerate(dist, start=1)])

# --- Run tests ---
g_results = g_test_per_bin(dist_a, dist_b)
chi2_results = chi2_test_per_bin(dist_a, dist_b)
data_a = expand_to_samples(dist_a)
data_b = expand_to_samples(dist_b)
ks_stat, ks_pval = ks_2samp(data_a, data_b)

# --- Save G-test, Chi-squared, and t-test results to Excel ---
g_results_df = pd.DataFrame(g_results, columns=["Bin", "G-statistic", "GS p-value"])
chi2_results_df = pd.DataFrame(chi2_results, columns=["Bin", "Chi2-statistic", "Chi2 p-value"])
ks_results_df = pd.DataFrame({"KS Statistic":[ks_stat], "KS p-value": [ks_pval]})

wbk = load_workbook(xl_file)

with pd.ExcelWriter(xl_file, engine="openpyxl",mode='a',if_sheet_exists='overlay') as writer:
    ks_results_df.to_excel(writer,sheet_name=xl_outsheet,startcol=1,startrow=1,index=False)
    g_results_df.to_excel(writer,sheet_name=xl_outsheet,startcol=1,startrow=5,index=False)
    chi2_results_df.to_excel(writer,sheet_name=xl_outsheet,startcol=4,startrow=5,index=False)

#g_results_df.to_csv("g_test_results.csv", index=False)
#chi2_results_df.to_csv("chi2_test_results.csv", index=False)

print("G-test, Chi-squared, and T-test results saved to Excel.")

# --- Plot and save CDFs ---
plt.figure(figsize=(10, 6))
data_a_sorted = np.sort(data_a)
data_b_sorted = np.sort(data_b)
cdf_a_vals = np.arange(1, len(data_a_sorted) + 1) / len(data_a_sorted)
cdf_b_vals = np.arange(1, len(data_b_sorted) + 1) / len(data_b_sorted)

plt.step(data_a_sorted, cdf_a_vals, where='post', label='Power Line Fires (Cumulative)', marker='o')
plt.step(data_b_sorted, cdf_b_vals, where='post', label='Large Non-PL Fires (Cumulative)', marker='s')

# Mark the maximum KS distance
all_points = np.union1d(data_a_sorted, data_b_sorted)
cdf_a_interp = np.searchsorted(data_a_sorted, all_points, side='right') / len(data_a_sorted)
cdf_b_interp = np.searchsorted(data_b_sorted, all_points, side='right') / len(data_b_sorted)
diffs = np.abs(cdf_a_interp - cdf_b_interp)
max_diff_idx = np.argmax(diffs)

plt.plot([all_points[max_diff_idx], all_points[max_diff_idx]],
         [cdf_a_interp[max_diff_idx], cdf_b_interp[max_diff_idx]],
         'k--', label=f'Max KS distance = {ks_stat:.3f}')

plt.xlabel('Wind Speed Bins (mph)')
plt.ylabel('Cumulative Probability')
plt.title('CDFs with KS Statistic Marked (Compressed Distributions)')
plt.legend()
plt.grid(axis='both', linestyle='--', alpha=0.5)
plt.tight_layout()
bin_edges = np.array(np.arange(xl_default['n']))  # Example bin edges
bin_shifted = bin_edges + 1
bin_centers = bin_edges - 0.5
#plt.bar(bin_centers, values, width=(bin_edges[1] - bin_edges[0]))
plt.xticks(bin_shifted, labels)

# Save the figure
plt.savefig("cdf_plot.png")
print("CDF plot saved as 'cdf_plot.png'.")
print("KS stat: " + str(ks_stat) + "   KS pval: " + str(ks_pval))
plt.show()
