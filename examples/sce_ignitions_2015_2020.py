import sys
from datetime import datetime
import pytz
import logging
from openpyxl import load_workbook
import xlrd  # pip install
import weather_config
import weather_utils

logging.basicConfig(level=weather_config.config['Default']['LOG_LEVEL'])

xl_data = weather_config.config['SCE']['XL_DATA_FILE']
weather_db = weather_config.config['SCE']['WEATHER_DB']
xl_input_sheet = weather_config.config['SCE']['XL_INPUT_SHEET']
xl_output_sheet = weather_config.config['SCE']['XL_OUTPUT_SHEET']
ttpl_raw = weather_config.config['SCE']['TIME_WINDOWS']
gtpl_raw = weather_config.config['SCE']['DISTANCE_WINDOWS']

tlst = ttpl_raw.split(',')
ttpl = (int(tlst[0]))
glst = gtpl_raw.split(',')
gtpl = (int(glst[0]))

try: 
    sceigndb = weather_utils.WeatherDB(weather_db)   
except:
    sceigndb = weather_utils.WeatherDB.create(weather_db)

logging.info('Opening workbook ' + xl_data)
wbk = load_workbook(filename=xl_data)

sht_all = wbk[xl_input_sheet]

try:
    sht_wind = wbk[xl_output_sheet]
except KeyError:
    wbk.create_sheet(xl_output_sheet)
    sht_wind = wbk[xl_output_sheet]
    
frow = int(weather_config.config['SCE']['FIRST_ROW'])
lrow = int(weather_config.config['SCE']['LAST_ROW'])


for irow in range(frow,lrow+1):
    
    # Copy row

    srow = str(irow)
    logging.info("Processing line " + srow)
    for icol in range(len(sht_all[srow])):
        sht_wind.cell(row=irow,column=icol+1).value = sht_all[irow][icol].value

    # Convert times to UTC for synoptic run. PG&E has multiple date/time formats
    # in their data set:
    # Case 1   B: Excel datetime    C: Excel datetime
    # Case 2   B: Excel date        C: Excel datetime
    # Case 3   B: Excel datetime    C: Excel time
    # Case 4   B: Excel date        C: Excel time
    # Case 5   B: Excel datetime    C: HH:MM
    # Case 6   B: DD/MM/YYYY        C: HH:MM  - Adding this for additional fires

    tcelld = "D"+ srow
    tcellt = "F"+ srow
    tmxld = sht_wind[tcelld].value
    tmxlt = sht_wind[tcellt].value

    is_xl_format = False
    tmxl = 0

    try:
        # For SCE data, dates are datetimes with incorrect time of 0:00

        tmpydt = tmxld.replace(hour=tmxlt.hour,minute=tmxlt.minute)
        
    except TypeError:
        logging.warning("Exiting on error. Saving workbook " + xl_data)
        wbk.save(xl_data)
        raise

    tmlocal = pytz.timezone("America/Los_Angeles").localize(tmpydt)
    tmutc = tmlocal.astimezone(pytz.utc)

    zigtime = weather_utils.TimeUtils(tmutc)
    lat = sht_wind["G"+srow].value
    lon = sht_wind["H"+srow].value

    # Get list of max wind data based on time and space windows
    # Output is m x n list of  [station_id,time,distance,maximum gust,count]

    ttpl = (ttpl,) if isinstance(ttpl,int) else ttpl
    gtpl = (gtpl,) if isinstance(gtpl,int) else gtpl
    
    try:
        max_gusts = weather_utils.get_max_gust(lat,lon,zigtime,ttpl,gtpl,sceigndb)
    except:
        logging.warning("Exiting on error. Saving workbook " + xl_data)
        wbk.save(xl_data)
        raise
    
    # Write output to cells
    
    icell = int(weather_config.config['SCE']['FREE_CELL'])
    for it in max_gusts:
        for ig in it:
            for val in ig:
                sht_wind.cell(row=irow,column=icell).value = val
                icell += 1

logging.info("Complete. Saving workbook " + xl_data)
wbk.save(xl_data)
