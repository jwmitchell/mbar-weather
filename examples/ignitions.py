import argparse
import datetime
import sys
from datetime import timedelta
import pytz
import logging
from openpyxl import load_workbook
import xlrd  # pip install

parse = argparse.ArgumentParser()
parse.add_argument('-u','--utility',choices=['PGE','SCE','SDGE'],required=True,help='utility=PGE,SCE,SDGE')
parse.add_argument('-f','--file',help='file=configuration file')
program_args=parse.parse_args()
import weather_config

if program_args.file is not None:
    weather_config.init(program_args.file)

import weather_utils

logging.basicConfig(level=weather_config.config['Default']['LOG_LEVEL'])

xl_data = weather_config.config[program_args.utility]['XL_DATA_FILE']
weather_db = weather_config.config[program_args.utility]['WEATHER_DB']
xl_input_sheet = weather_config.config[program_args.utility]['XL_INPUT_SHEET']
xl_output_sheet = weather_config.config[program_args.utility]['XL_OUTPUT_SHEET']
xl_date_col = weather_config.config[program_args.utility]['XL_DATE_COLUMN']
xl_time_col = weather_config.config[program_args.utility]['XL_TIME_COLUMN']
xl_lat_col = weather_config.config[program_args.utility]['XL_LAT_COLUMN']
xl_long_col = weather_config.config[program_args.utility]['XL_LONG_COLUMN']
ttpl_raw = weather_config.config[program_args.utility]['TIME_WINDOWS']
gtpl_raw = weather_config.config[program_args.utility]['DISTANCE_WINDOWS']


tlst = ttpl_raw.split(',')
ttpl = (int(tlst[0]),)
ii = 0
for itt in tlst:
    if ii>0:
        ttpl = ttpl + (int(tlst[ii]),)
    ii+=1

glst = gtpl_raw.split(',')
gtpl = (int(glst[0]),)
ii = 0
for igt in glst:
    if ii>0:
        gtpl = gtpl + (int(glst[ii]),)
    ii+=1

try: 
    igndb = weather_utils.WeatherDB(weather_db)   
except:
    igndb = weather_utils.WeatherDB.create(weather_db)

logging.info('Opening workbook ' + xl_data)
wbk = load_workbook(filename=xl_data)

sht_all = wbk[xl_input_sheet]

try:
    sht_wind = wbk[xl_output_sheet]
except KeyError:
    sht_wind = wbk.copy_worksheet(sht_all)
    sht_wind.title = xl_output_sheet
    
frow = int(weather_config.config[program_args.utility]['FIRST_ROW'])
lrow = int(weather_config.config[program_args.utility]['LAST_ROW'])


for irow in range(frow,lrow+1):
    
    # Copy row

    srow = str(irow)
    logging.info("Processing line " + srow)
    for icol in range(len(sht_all[srow])):
        sht_wind.cell(row=irow,column=icol+1).value = sht_all[irow][icol].value

    # Convert times to UTC for synoptic run. PG&E has multiple date/time formats
    # in their data set:
    # Case 1   B: Excel datetime    C: Excel datetime
    # Case 2   B: Excel date        C: Excel datetime
    # Case 3   B: Excel datetime    C: Excel time
    # Case 4   B: Excel date        C: Excel time
    # Case 5   B: Excel datetime    C: HH:MM
    # Case 6   B: DD/MM/YYYY        C: HH:MM  - Adding this for additional fires

    tcelld = weather_config.config[program_args.utility]['XL_DATE_COLUMN'] + srow
    tcellt = weather_config.config[program_args.utility]['XL_TIME_COLUMN'] + srow
    tmxld = sht_wind[tcelld].value
    tmxlt = sht_wind[tcellt].value

    is_xl_format = False
    tmxl = 0

    try:
        if type(tmxld) is str:
            if '/' in tmxld:
                dd = tmxld.split('/')
                tt = tmxlt.split(':')
                tmpydt = datetime.datetime(int(dd[0]),int(dd[1]),int(dd[2]),\
                                           int(tt[0]),int(tt[1]))
        if isinstance(tmxld,datetime.datetime):
            if isinstance(tmxlt,datetime.time):
                tmpydt = datetime.datetime.combine(tmxld,tmxlt)
            elif tmxld.hour==0 and tmxld.minute==0:
                if type(tmxlt) is str:
                    tt = tmxlt.split(':')
                    tmpydt = tmxld
                    tmpydt = tmpydt.replace(hour=int(tt[0]),minute=int(tt[1]))
                else:
                    raise ValueError("Unknown time data type" + type(tmxlt))
            else:
                tmpydt = tmxld
        elif float(tmxld) > 40000.00 and float(tmxld) < 50000.00:
            is_xl_format = True
            if not float(tmxld).is_integer():
                tmxl = float(tmxld) 
            elif float(tmxlt) > 40000.0 and float(tmxlt) < 50000.0:
                tmxl = float(tmxlt)
            elif float(tmxlt) < 1.0:
                tmxl = float(tmxld) + float(tmxlt)
            else:
                raise TypeError("Unrecognized time value pair B:" + str(tmxld) + " C:" + str(tmxlt))
            tmpydt = datetime(*xlrd.xldate_as_tuple(tmxl,0))

        # For SCE data, dates are datetimes with incorrect time of 0:00

        # tmpydt = tmxld.replace(hour=tmxlt.hour,minute=tmxlt.minute)
        
    except TypeError:
        logging.warning("Exiting on error. Saving workbook " + xl_data)
        wbk.save(xl_data)
        raise

    tmlocal = pytz.timezone("America/Los_Angeles").localize(tmpydt)
    tmutc = tmlocal.astimezone(pytz.utc)

    zigtime = weather_utils.TimeUtils(tmutc)
    lat = sht_wind[weather_config.config[program_args.utility]['XL_LAT_COLUMN']+srow].value
    lon = sht_wind[weather_config.config[program_args.utility]['XL_LONG_COLUMN']+srow].value

    # Get list of max wind data based on time and space windows
    # Output is m x n list of  [station_id,time,distance,maximum gust,count]

    ttpl = (ttpl,) if isinstance(ttpl,int) else ttpl
    gtpl = (gtpl,) if isinstance(gtpl,int) else gtpl
    
    try:
        max_gusts = weather_utils.get_max_gust(lat,lon,zigtime,ttpl,gtpl,igndb)
    except:
        logging.warning("Exiting on error. Saving workbook " + xl_data)
        wbk.save(xl_data)
        raise
    
    # Write output to cells
    
    icell = int(weather_config.config[program_args.utility]['FREE_CELL'])
    for it in max_gusts:
        for ig in it:
            for val in ig:
                sht_wind.cell(row=irow,column=icell).value = val
                icell += 1

logging.info("Complete. Saving workbook " + xl_data)
wbk.save(xl_data)
