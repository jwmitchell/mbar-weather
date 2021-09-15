#### Copyright 2020, M-bar Technologies and Consulting, LLC #### 
#### Distributed under the Gnu General Public Licence, v3   ####
#
# This program is designed to provide a random control sample of weather data
# to accompany analyis done in the ignitions.py or circuit_damage.py script.
#
# The routine randomly draws an ignition event and then picks a random time
# specified in the *_mc.ini configuration file. Maximum wind
# speed for the geographic location of the random ignition event is determined,
# and values are written out to a new sheet in the excel file. 
# 

import argparse
import sys
from datetime import timedelta
import pytz
import logging
from openpyxl import load_workbook
import xlrd  # pip install
import random

parse = argparse.ArgumentParser()
parse.add_argument('-u','--utility',choices=['PGE','SCE','SDGE'],required=True,help='utility=PGE,SCE,SDGE')
parse.add_argument('-f','--file',help='file=configuration file')
program_args=parse.parse_args()
import weather_config

if program_args.file is not None:
    weather_config.init(program_args.file)

import weather_utils

logging.basicConfig(level=weather_config.config['Default']['LOG_LEVEL'])

xl_data = weather_config.config[program_args.utility]['XL_DATA_FILE']
weather_db = weather_config.config[program_args.utility]['WEATHER_DB']
xl_input_sheet = weather_config.config[program_args.utility]['XL_INPUT_SHEET']
xl_output_sheet = weather_config.config[program_args.utility]['XL_OUTPUT_SHEET']
xl_date_col = weather_config.config[program_args.utility]['XL_DATE_COLUMN']
xl_time_col = weather_config.config[program_args.utility]['XL_TIME_COLUMN']
xl_lat_col = weather_config.config[program_args.utility]['XL_LAT_COLUMN']
xl_long_col = weather_config.config[program_args.utility]['XL_LONG_COLUMN']
ttpl_raw = weather_config.config[program_args.utility]['TIME_WINDOWS']
gtpl_raw = weather_config.config[program_args.utility]['DISTANCE_WINDOWS']
nevents = int(weather_config.config[program_args.utility]['EVENTS'])
frow = int(weather_config.config[program_args.utility]['FIRST_ROW'])
lrow = int(weather_config.config[program_args.utility]['LAST_ROW'])
offset = int(weather_config.config[program_args.utility]['ROW_OFFSET']) or 0
start_date = weather_config.config[program_args.utility]['START_DATE']
end_date = weather_config.config[program_args.utility]['END_DATE']

tlst = ttpl_raw.split(',')
ttpl = (int(tlst[0]),)
ii = 0
for itt in tlst:
    if ii>0:
        ttpl = ttpl + (int(tlst[ii]),)
    ii+=1

glst = gtpl_raw.split(',')
gtpl = (int(glst[0]),)
ii = 0
for igt in glst:
    if ii>0:
        gtpl = gtpl + (int(glst[ii]),)
    ii+=1

try: 
    igndb = weather_utils.WeatherDB(weather_db)   
except:
    igndb = weather_utils.WeatherDB.create(weather_db)

logging.info('Opening workbook ' + xl_data)
wbk = load_workbook(filename=xl_data)

sht_all = wbk[xl_input_sheet]
try:
    sht_wind = wbk[xl_output_sheet]
except KeyError:
    sht_wind = wbk.copy_worksheet(sht_all)
    sht_wind.title = xl_output_sheet

dtstart = weather_utils.TimeUtils(start_date)
dtend = weather_utils.TimeUtils(end_date)

for ievt in range(nevents):

    irow = frow + random.randrange(lrow-frow)
    
    # Copy row

    srow = str(irow)
    logging.info("Event #" + str(ievt) + " Processing line " + srow)
    jrow = ievt + frow + offset
    logging.debug("Writing row " + str(jrow))
    for icol in range(len(sht_all[irow])):
        sht_wind.cell(row=jrow,column=icol+1).value = sht_all[irow][icol].value

    tcelld = weather_config.config[program_args.utility]['XL_DATE_COLUMN'] + srow
    tcellt = weather_config.config[program_args.utility]['XL_TIME_COLUMN'] + srow
    
    zigtime = weather_utils.TimeUtils.randtime(dtstart,dtend)

    lat = sht_wind[weather_config.config[program_args.utility]['XL_LAT_COLUMN']+srow].value
    lon = sht_wind[weather_config.config[program_args.utility]['XL_LONG_COLUMN']+srow].value

    logging.debug("time, lat, lon = " + str(zigtime.datetime) + " " + str(lat) + " " + str(lon))

    # Get list of max wind data based on time and space windows
    # Output is m x n list of  [station_id,time,distance,maximum gust,count]

    ttpl = (ttpl,) if isinstance(ttpl,int) else ttpl
    gtpl = (gtpl,) if isinstance(gtpl,int) else gtpl
    
    try:
        max_gusts = weather_utils.get_max_gust(lat,lon,zigtime,ttpl,gtpl,igndb)
    except:
        logging.warning("Exiting on error. Saving workbook " + xl_data)
        wbk.save(xl_data)
        raise

    first_wg_cell = int(weather_config.config[program_args.utility]['FREE_CELL'])
    
    # Clear wind gust cells
    
    nmgcells = len(gtpl)*len(ttpl)*5
    for ic in range(nmgcells):
        sht_wind.cell(row=jrow, column=first_wg_cell+ic).value = None


    # Write output to cells
                
    icell = first_wg_cell
    for it in max_gusts:
        for ig in it:
            for val in ig:
                sht_wind.cell(row=jrow,column=icell).value = val
                icell += 1
    
logging.info("Complete. Saving workbook " + xl_data)
wbk.save(xl_data)
